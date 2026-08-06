"""Excel report exporter."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from backend.app.reporting.enums import DocumentNodeType, ExportFormat
from backend.app.reporting.models import DocumentNode, ExportResult, RenderedDocument


@dataclass(frozen=True, slots=True)
class ExcelExporter:
    """Export rendered documents as Excel workbooks."""

    format: ExportFormat = ExportFormat.EXCEL

    def export(self, document: RenderedDocument) -> ExportResult:
        try:
            from openpyxl import Workbook  # type: ignore[import-untyped]
        except ImportError as exc:
            msg = "openpyxl is required for Excel export"
            raise RuntimeError(msg) from exc

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["metric", "value"])
        stats = document.statistics
        for key, value in {
            "total_devices": stats.total_devices,
            "reachable_devices": stats.reachable_devices,
            "unreachable_devices": stats.unreachable_devices,
            "discovery_success_pct": stats.discovery_success_pct,
            "netbox_accuracy_pct": stats.netbox_accuracy_pct,
            "compliance_score": stats.compliance_score,
            "critical_findings": stats.critical_findings,
            "major_findings": stats.major_findings,
            "minor_findings": stats.minor_findings,
            "missing_devices": stats.missing_devices,
            "extra_devices": stats.extra_devices,
            "changed_devices": stats.changed_devices,
            "interface_changes": stats.interface_changes,
            "vlan_changes": stats.vlan_changes,
            "configuration_changes": stats.configuration_changes,
        }.items():
            summary.append([key, value])

        tables = _collect_tables(document.root)
        for index, (title, headers, rows) in enumerate(tables, start=1):
            sheet_name = _sheet_name(title, index)
            sheet = workbook.create_sheet(title=sheet_name)
            if headers:
                sheet.append([_serialize_value(item) for item in headers])
            for row in rows:
                sheet.append([_serialize_value(item) for item in row])

        recommendations = workbook.create_sheet(title="Recommendations")
        recommendations.append(
            [
                "recommendation_id",
                "category",
                "action",
                "priority",
                "subject_type",
                "subject_id",
                "reason_code",
            ]
        )
        for item in document.recommendations:
            recommendations.append(
                [
                    item.recommendation_id,
                    item.category.value,
                    item.action.value,
                    item.priority.value,
                    item.subject_type,
                    item.subject_id,
                    item.reason_code,
                ]
            )

        buffer = BytesIO()
        workbook.save(buffer)
        filename = f"{document.report_type.value}.xlsx"
        return ExportResult(
            format=self.format,
            content=buffer.getvalue(),
            mime_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            filename=filename,
        )


def _collect_tables(
    node: DocumentNode,
    *,
    section: str = "document",
) -> list[tuple[str, tuple[object, ...], tuple[tuple[object, ...], ...]]]:
    tables: list[tuple[str, tuple[object, ...], tuple[tuple[object, ...], ...]]] = []
    current_section = section
    if node.node_type == DocumentNodeType.SECTION:
        current_section = str(node.attributes.get("title", section))

    if node.node_type == DocumentNodeType.TABLE:
        headers_value = node.attributes.get("headers", ())
        rows_value = node.attributes.get("rows", ())
        headers: tuple[object, ...] = (
            tuple(headers_value) if isinstance(headers_value, tuple) else tuple()
        )
        rows: tuple[tuple[object, ...], ...] = (
            tuple(tuple(row) for row in rows_value)
            if isinstance(rows_value, tuple)
            else tuple()
        )
        tables.append((current_section, headers, rows))

    for child in node.children:
        tables.extend(_collect_tables(child, section=current_section))
    return tables


def _serialize_value(value: object) -> object:
    if isinstance(value, (tuple, list)):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return ", ".join(f"{key}={item}" for key, item in value.items())
    return value


def _sheet_name(title: str, index: int) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in title)
    cleaned = cleaned.strip("_")[:25] or f"Sheet{index}"
    return cleaned
